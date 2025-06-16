"""
Обработчик Excel файлов для работы с адресами
"""
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl
from utils.logger import setup_logger
import config

logger = setup_logger()

class ExcelProcessor:
    """Обработчик Excel файлов"""
    
    def __init__(self):
        """Инициализация процессора"""
        self.workbook = None
        self.worksheet = None
    
    def load_workbook(self):
        """
        Загружает Excel файл для обработки
        
        Returns:
            tuple: (workbook, worksheet)
        """
        if not Path(config.INPUT_EXCEL).exists():
            raise FileNotFoundError(f"Файл {config.INPUT_EXCEL} не найден")
        
        # Проверяем, не открыт ли файл
        try:
            # Пытаемся открыть файл для записи
            with open(config.INPUT_EXCEL, 'r+b'):
                pass
        except PermissionError:
            raise PermissionError(f"Файл {config.INPUT_EXCEL} открыт в другом приложении. Закройте Excel.")
        
        logger.info(f"Загружаем файл: {config.INPUT_EXCEL}")
        
        try:
            self.workbook = openpyxl.load_workbook(config.INPUT_EXCEL)
            self.worksheet = self.workbook.active
            
            # Определяем количество строк с данными
            max_row = self.worksheet.max_row
            data_rows = max_row - config.EXCEL_START_ROW
            
            logger.info(f"Файл загружен успешно. Строк для обработки: {data_rows}")
            
            return self.workbook, self.worksheet
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке Excel файла: {e}")
            raise
    
    def create_backup(self):
        """Создаёт резервную копию исходного файла"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"addresses_backup_{timestamp}.xlsx"
            backup_path = Path(config.BACKUP_DIR) / backup_name
            
            shutil.copy2(config.INPUT_EXCEL, backup_path)
            logger.info(f"Создана резервная копия: {backup_path}")
            
        except Exception as e:
            logger.warning(f"Не удалось создать резервную копию: {e}")
    
    def read_address_row(self, row_num):
        """
        Читает адресные данные из строки Excel
        
        Args:
            row_num (int): Номер строки (1-based)
            
        Returns:
            dict: Данные адреса или None если строка пустая
        """
        try:
            settlement = self.worksheet[f"{config.EXCEL_SETTLEMENT_COL}{row_num}"].value
            street = self.worksheet[f"{config.EXCEL_STREET_COL}{row_num}"].value  
            house = self.worksheet[f"{config.EXCEL_HOUSE_COL}{row_num}"].value
            
            # Проверяем есть ли данные
            if not any([settlement, street, house]):
                return None
            
            # Очищаем данные
            address_data = {
                'row_num': row_num,
                'settlement': str(settlement).strip() if settlement else '',
                'street': str(street).strip() if street else '',
                'house': str(house).strip() if house else ''
            }
            
            # Проверяем, что есть хотя бы населённый пункт
            if not address_data['settlement']:
                return None
                
            return address_data
            
        except Exception as e:
            logger.debug(f"Ошибка при чтении строки {row_num}: {e}")
            return None
    
    def write_result(self, row_num, status, details=''):
        """
        Записывает результат проверки в Excel
        
        Args:
            row_num (int): Номер строки
            status (str): Статус проверки (Да/Проверить/Нет)
            details (str): Дополнительная информация
        """
        try:
            self.worksheet[f"{config.EXCEL_RESULT_COL}{row_num}"] = status
            self.worksheet[f"{config.EXCEL_DETAILS_COL}{row_num}"] = details
            
        except Exception as e:
            logger.error(f"Ошибка записи результата в строку {row_num}: {e}")
    
    def save_results(self, results):
        """
        Сохраняет все результаты в Excel файл
        
        Args:
            results (list): Список результатов проверки
        """
        try:
            # Создаём резервную копию перед сохранением
            self.create_backup()
            
            # Записываем результаты
            for result in results:
                if result:  # Проверяем что результат не None
                    self.write_result(
                        result['row_num'],
                        result['status'], 
                        result.get('details', '')
                    )
            
            # Сохраняем файл
            self.workbook.save(config.OUTPUT_EXCEL)
            logger.info(f"Результаты сохранены: {config.OUTPUT_EXCEL}")
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении результатов: {e}")
            raise
    
    def get_total_rows(self):
        """
        Возвращает общее количество строк для обработки
        
        Returns:
            int: Количество строк
        """
        if not self.worksheet:
            return 0
        
        return max(0, self.worksheet.max_row - config.EXCEL_START_ROW)
    
    def close(self):
        """Закрывает рабочую книгу"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None